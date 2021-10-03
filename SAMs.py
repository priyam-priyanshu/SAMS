from tkinter import *
from tkinter import filedialog
import openpyxl
import tkinter.messagebox as tmsg
import smtplib

root = Tk()
# GEOMETRY
root.geometry(f"{root.winfo_screenwidth()}x{root.winfo_screenheight()}")
root.title("SAMs")
root.configure(bg="#B1D4E0")  # Background colour
root.wm_iconbitmap(r"temp\icon.ico")

# DEF FUNCTIONS
def pprint(txt):
    sep = "-"*75
    txt = txt + f"\n{sep}"
    Label(out_box, text=txt)
    pass

file = ""
def browse_file():
    global file
    file = filedialog.askopenfilename(title="Select the Excel Sheet", filetypes=(("Excel", "*.xls*"), ("Office Excel", "*.xlsx*")))
    if len(str(file)) > 0:
        generate_btn["state"] = "normal"
        status.configure(fg="#1A2421", text="File Added!")
        status.place(relx=0.3, rely=0.95)
    return file

def send_mail(to, content, step, maxx):
    try:
        s = smtplib.SMTP(f'smtp.{radio_var.get()}.com', 587)
        s.starttls()
        s.login(sender_email_var.get(), sender_password_var.get())
        message = "\n" + content
        s.sendmail(sender_email_var.get(), to, message)

        if step == maxx:
            s.quit()

    except Exception as e:
        tmsg.showinfo("Mail Error", "Kindly check the login credentials.\n\nIf using GMail, we are counted among the less secure apps and for some users our services might not be available.\n\nWe are sorry for the inconvinence caused.")
    pass

def generte_mail():
    txt = txt_box.get(1.0, "end-1c")
    sep = "\n" + "-" * 75 + "\n"

    try:
        work_book = openpyxl.load_workbook(file)
        sheet = work_book.active
        max_row = sheet.max_row
        new_txt = ""

        for i in range(2, max_row + 1):
            name_cell = sheet.cell(row=i, column=name_column.get()).value
            new_txt = txt.replace(name_var.get(), name_cell)
            email_cell = sheet.cell(row=i, column=email_column.get()).value

            if len(email_var.get()) > 0:
                new_txt = new_txt.replace(email_var.get(), email_cell)
            if var_1_column.get() > 0:
                var_1_sheet = sheet.cell(row=i, column=var_1_column.get()).value
                new_txt = new_txt.replace(var_1.get(), var_1_sheet)
            if var_2_column.get() > 0:
                var_2_sheet = sheet.cell(row=i, column=var_2_column.get()).value
                new_txt = new_txt.replace(var_2.get(), var_2_sheet)
            if var_3_column.get() > 0:
                var_3_sheet = sheet.cell(row=i, column=var_3_column.get()).value
                new_txt = new_txt.replace(var_3.get(), var_3_sheet)

            new_txt = f"SUBJECT: {subject_var.get()}\n\n{new_txt}"

            out_box.insert(INSERT, new_txt + sep)

            if i == 2:
                res = tmsg.askyesnocancel("Confimation Required",
                                          "Kindly confim the below generated mail.\nHit yes if confirmed!")
                if not res:
                    exit()

            send_mail(email_cell, new_txt, i, max_row)

            if i == max_row+1:
                tmsg.showinfo("Congratulations!", f"All {max_row-1} mails sent successfully !")
    except Exception as e:
        out_box.insert(INSERT, "Excel Sheet error" + sep)
        tmsg.showinfo("Excel Sheet Error", "Kindly check the column values provided are correct.")

    pass



# DEF HELP


# MENUBAR
main_menu = Menu(root)
file_menu = Menu(main_menu, tearoff=0, bg="#D4F1F4")
file_menu.add_command(label="Browse Sheet", command=browse_file)
# file_menu.add_command(label="Browse Text", command=browse_file) ##########################################
file_menu.add_separator()
file_menu.add_command(label="Exit", command=exit)
root.config(menu=main_menu)
main_menu.add_cascade(label="File", menu=file_menu)

help_menu = Menu(main_menu, tearoff=0, bg="#D4F1F4")
help_menu.add_command(label="About Us")
help_menu.add_command(label="Contact Us")
root.config(menu=main_menu)
main_menu.add_cascade(label="Help", menu=help_menu)

# VARIABLES
name_var = StringVar()
name_column = IntVar()
email_var = StringVar()
subject_var = StringVar()
email_column = IntVar()
var_1 = StringVar()
var_1_column = IntVar()
var_2 = StringVar()
var_2_column = IntVar()
var_3 = StringVar()
var_3_column = IntVar()

sender_email_var = StringVar()
sender_password_var = StringVar()
radio_var = StringVar()
radio_var.set("outlook")

# VAR_BOX
f1 = Frame(root, bg="#145DA0")
f1.place(relx=0.005, rely=0.01, width=root.winfo_screenwidth()/3.5, height=root.winfo_screenheight()-60)
Label(f1, text="VARIABLE   BOX", font="Harrington 17 bold", bg="#145DA0", fg="white").place(relx=0.2, rely=0.01)
Label(f1, text="---------------", bg="white", fg="white").place(relx=0.005, rely=0.05, height=0.3, width=root.winfo_screenwidth()/3.5)

Label(f1, text="NAME", font="courier 9 bold", bg="#145DA0", fg="white").place(relx=0.03, rely=0.09)
Label(f1, text="VALUE", font="courier 9 bold", bg="#145DA0", fg="white").place(relx=0.4, rely=0.09)
Label(f1, text="COLUMN", font="courier 9 bold", bg="#145DA0", fg="white").place(relx=0.8, rely=0.09)

dist = 0.08

Label(f1, text="Name:", font="courier 17 bold", bg="#145DA0", fg="white").place(relx=0.1, rely=0.13)
Entry(f1, textvariable=name_var, font="courier 17 bold").place(relx=0.04+0.23, rely=0.14, width=230, height=25)
Entry(f1, textvariable=name_column, font="courier 17 bold").place(relx=0.04+0.8, rely=0.14, width=30, height=25)

Label(f1, text="EMail:", font="courier 17 bold", bg="#145DA0", fg="white").place(relx=0.04+0.03, rely=0.13+dist)
Entry(f1, textvariable=email_var, font="courier 17 bold").place(relx=0.04+0.23, rely=0.14+(1*dist), width=230, height=25)
Entry(f1, textvariable=email_column, font="courier 17 bold").place(relx=0.04+0.8, rely=0.14+(1*dist), width=30, height=25)

Label(f1, text="Subject:", font="courier 17 bold", bg="#145DA0", fg="white").place(relx=0.013, rely=0.13+(2*dist))
Entry(f1, textvariable=subject_var, font="courier 17 bold").place(relx=0.04+0.23, rely=0.14+(2*dist), width=230, height=25)
Entry(f1, font="courier 17 bold").place(relx=0.04+0.8, rely=0.14+(2*dist), width=30, height=25)

Label(f1, text="Var 1:", font="courier 17 bold", bg="#145DA0", fg="white").place(relx=0.04+0.03, rely=0.13+(3*dist))
Entry(f1, textvariable=var_1, font="courier 17 bold").place(relx=0.04+0.23, rely=0.14+(3*dist), width=230, height=25)
Entry(f1, textvariable=var_1_column, font="courier 17 bold").place(relx=0.04+0.8, rely=0.14+(3*dist), width=30, height=25)

Label(f1, text="Var 2:", font="courier 17 bold", bg="#145DA0", fg="white").place(relx=0.04+0.03, rely=0.13+(4*dist))
Entry(f1, textvariable=var_2, font="courier 17 bold").place(relx=0.04+0.23, rely=0.14+(4*dist), width=230, height=25)
Entry(f1, textvariable=var_2_column, font="courier 17 bold").place(relx=0.04+0.8, rely=0.14+(4*dist), width=30, height=25)

Label(f1, text="Var 3:", font="courier 17 bold", bg="#145DA0", fg="white").place(relx=0.04+0.03, rely=0.13+(5*dist))
Entry(f1, textvariable=var_3, font="courier 17 bold").place(relx=0.04+0.23, rely=0.14+(5*dist), width=230, height=25)
Entry(f1, textvariable=var_3_column, font="courier 17 bold").place(relx=0.04+0.8, rely=0.14+(5*dist), width=30, height=25)

Label(f1, text="   ", font="courier 17 bold", bg="#B1D4E0").place(relx=0, rely=0.64, width=450, height=7)

# Login Box
Label(f1, text="SENDER's CREDENTIALS", bg="#145DA0", fg="white", font="Harrington 17 bold").place(relx=0.12, rely=0.66)
Label(f1, text="  ", bg="white").place(relx=0, rely=0.71, width=450, height=1)

Label(f1, text="EMail:", font="courier 17 bold", bg="#145DA0", fg="white").place(relx=0.06, rely=0.74)
Entry(f1, textvariable=sender_email_var, font="courier 17 bold").place(relx=0.4, rely=0.75, width=230, height=25)

Label(f1, text="Password:", font="courier 17 bold", bg="#145DA0", fg="white").place(relx=0.06, rely=0.82)
Entry(f1, textvariable=sender_password_var, font="courier 17 bold", show="*").place(relx=0.4, rely=0.83, width=230, height=25)

radio = Radiobutton(f1, text="Outlook", variable=radio_var, value="outlook", bg="#145DA0", font="courier 17 bold", fg="#0C2D48", activebackground="#145DA0", activeforeground="#0C2D48")
radio.place(relx=0.05, rely=0.89)

radio = Radiobutton(f1, text="GMail", variable=radio_var, value="gmail", bg="#145DA0", font="courier 17 bold", fg="#0C2D48", activebackground="#145DA0", activeforeground="#0C2D48")
radio.place(relx=0.5, rely=0.89)

status = Label(f1, text="Choose an Excel Sheet...", bg="#145DA0", fg="#7C0A02", font="courier 15 bold")
status.place(relx=0.15, rely=0.95)

# TEXT BOX
scroll = Scrollbar(root)
scroll.place(relx=0.984, rely=0.01, width=15, height=515)
txt_box = Text(root, bg="#2E8BC0", fg="white", font=("comic sans ms", 19, "bold"), border=0, yscrollcommand=scroll.set, wrap=WORD)
txt_box.place(relx=0.296, rely=0.01, width=1055, height=root.winfo_screenheight()/1.68)
scroll.configure(command=txt_box.yview)


# OUTPUT BOX
f3 = Frame(root, bg="#0C2D48")
f3.place(relx=0.296, rely=0.643, width=1070, height=287)

Label(f3, text="OUTPUT", font="Harrington 17 bold", bg="#0C2D48", fg="white").pack()

scroll_2 = Scrollbar(f3, bg="#0C2D48", troughcolor="#2E8BC0")
scroll_2.place(relx=0.987, rely=0, width=15, height=287)
out_box = Text(f3, bg="#0C2D48", fg="#FC7676", font="courier 14 bold ", yscrollcommand=scroll_2.set, wrap=WORD)
out_box.place(relx=0.01, rely=0.13, width=900, height=240)
scroll_2.configure(command=out_box.yview)

Button(f3, text="Browse", command=browse_file, font="courier 15 bold").place(relx=0.88, rely=0.6)
generate_btn = Button(f3, text="Generate", command=generte_mail, font="courier 15 bold", state="disabled")
generate_btn.place(relx=0.87, rely=0.8)

root.mainloop()